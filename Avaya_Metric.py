from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


# from openpyxl.utils import datetime_to_excel
def can_convert_to_int(value):
    try:
        int(value)
        return True
    except ValueError:
        return False

def convert_to_date(value):
    if can_convert_to_int(value):
        duration = int(value)
        hours, remainder = divmod(duration, 3600)
        minutes, seconds = divmod(remainder, 60)
        formatted_time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        value = str(formatted_time)
    return value

class Agent:
    def __int__(self):
        self.Agent_Name = None
        self.ACD_Calls = None
        self.Avg_ACD_Time = None
        self.Avg_ACW_Time = None
        self.Agent_Occupancy_ACW = None
        self.Agent_Occupancy_no_ACW = None
        self.Extn_In_Calls = None
        self.Avg_Extn_In_Time = None
        self.Extn_Out_Calls = None
        self.Avg_Extn_Out_Time = None
        self.ACD_Time = None
        self.ACW_Time = None
        self.Agent_Ring_Time = None
        self.Other_Time = None
        self.AUX_Time = None
        self.Avail_Time = None
        self.Staffed_Time = None


path = "/Users/jasonchoi/PycharmProjects/Py+Excel/March Raw.xlsx"
# path = "C:\\Users\\jasonchoi\\PycharmProjects\\Py+Excel\\demo.xlsx"
# ws = Workbook.load_workbook(path)
wb = load_workbook(path)
sheet = wb.active
Agent_List = []
for i in range(1, sheet.max_row + 1):
    Curr_Emp = Agent()
    for j in range(1, sheet.max_column + 1):
        cell_obj = sheet.cell(row=i, column=j)
        cell = 0 if cell_obj.value is None else cell_obj.value

        if j == 1 :
            Curr_Emp.Agent_Name = cell
        if j == 2:
            Curr_Emp.ACD_Calls = cell
        if j == 3:
            Curr_Emp.Avg_ACD_Time = convert_to_date(cell)
        if j == 4:
            Curr_Emp.Avg_ACW_Time = convert_to_date(cell)
        if j == 5:
            Curr_Emp.Agent_Occupancy_ACW = cell
        if j == 6:
            Curr_Emp.Agent_Occupancy_no_ACW = cell
        if j == 7:
            Curr_Emp.Extn_In_Calls = cell
        if j == 8:
            Curr_Emp.Avg_Extn_In_Time = cell
        if j == 9:
            Curr_Emp.Extn_Out_Calls = cell
        if j == 10:
            Curr_Emp.Avg_Extn_Out_Time = cell
        if j == 11:
            Curr_Emp.ACD_Time = convert_to_date(cell)
        if j == 12:
            Curr_Emp.ACW_Time = convert_to_date(cell)
        if j == 13:
            Curr_Emp.Agent_Ring_Time = convert_to_date(cell)
        if j == 14:
            Curr_Emp.Other_Time = convert_to_date(cell)
        if j == 15:
            Curr_Emp.AUX_Time = convert_to_date(cell)
        if j == 16:
            Curr_Emp.Avail_Time = convert_to_date(cell)
        if j == 17:
            Curr_Emp.Staffed_Time = convert_to_date(cell)

    Agent_List.append(Curr_Emp)

metric = True

if metric:
    Metric_WB = Workbook()
    sheet = Metric_WB.active

    for i, agent in enumerate(Agent_List, start=1):
        sheet.cell(row=i, column=1).value = agent.Agent_Name
        sheet.cell(row=i, column=2).value = agent.ACD_Calls
        sheet.cell(row=i, column=3).value = agent.Avg_ACD_Time
        sheet.cell(row=i, column=4).value = agent.Avg_ACW_Time
        sheet.cell(row=i, column=5).value = agent.Agent_Occupancy_ACW
        sheet.cell(row=i, column=6).value = agent.Agent_Occupancy_no_ACW
        sheet.cell(row=i, column=7).value = agent.Extn_In_Calls
        sheet.cell(row=i, column=8).value = agent.Avg_Extn_In_Time
        sheet.cell(row=i, column=9).value = agent.Extn_Out_Calls
        sheet.cell(row=i, column=10).value = agent.Avg_Extn_Out_Time
        sheet.cell(row=i, column=11).value = agent.ACD_Time
        sheet.cell(row=i, column=12).value = agent.ACW_Time
        sheet.cell(row=i, column=13).value = agent.Agent_Ring_Time
        sheet.cell(row=i, column=14).value = agent.Other_Time
        sheet.cell(row=i, column=15).value = agent.AUX_Time
        sheet.cell(row=i, column=16).value = agent.Avail_Time
        sheet.cell(row=i, column=17).value = agent.Staffed_Time

    Metric_WB.save("example.xlsx")
